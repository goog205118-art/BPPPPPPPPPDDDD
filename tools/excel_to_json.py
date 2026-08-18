import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Missing Excel file path")

    file_path = Path(sys.argv[1])
    if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise SystemExit("Only .xlsx/.xlsm/.xltx/.xltm files are supported")

    records = []
    workbook = load_workbook(file_path, read_only=True, data_only=True)

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [normalize(cell) for cell in rows[0]]
        if not any(headers):
            continue

        for row in rows[1:]:
            record = {"来源工作表": sheet.title}
            has_value = False
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row[index] if index < len(row) else ""
                text = normalize(value)
                if text:
                    has_value = True
                record[header] = text
            if has_value:
                records.append(record)

    print(json.dumps(records, ensure_ascii=False))


if __name__ == "__main__":
    main()
